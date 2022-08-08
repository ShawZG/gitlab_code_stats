# -*- coding: utf-8 -*-
#
# Copyright (C) 2020 xiao zhiguo <xiaozhiguo@uniontech.com>
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Lesser General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Lesser General Public License for more details.
#
# You should have received a copy of the GNU Lesser General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.

import os
import argparse
import gitlab
from datetime import date, datetime
from openpyxl import Workbook
from gitlab_project_manager import GitLabProjectManager


class GitLabProjectStats:
    """
    Args:
        project(str): 指定统计的项目名称
        branch(str): 设置统计项目的分支
        since(str): 设置统计的开始时间，格式YYYY-MM-DD
        until(str): 设置统计的结束时间，不包括该时间，格式YYYY-MM-DD
    """
    def __init__(self, project: str = None, branch: str = None, since: str = None, until: str = None):
        self._gitlab_url: str = ""
        self._gitlab_access_token: str = ""

        self._filter_project_name: str = project
        self._filter_branch: str = branch
        self._since: str = since
        self._until: str = until

        # 需要统计的项目
        self._target_projects: list = []
        self._project_branches: list = []
        self._project_branch_commit_data: dict = {}

    def _get_gitlab_projects(self):
        """
        根据project名称，获取项目信息
        """
        gl = gitlab.Gitlab(url=self._gitlab_url, private_token=self._gitlab_access_token)
        gl.auth()
        gitlab_all_projects: list = gl.projects.list(all=True)
        print("================================================================================")
        for project in gitlab_all_projects:
            print("project id:", project.id, project.name, project.web_url)
        if self._filter_project_name:
            # 完全匹配
            self._target_projects = [project for project in gitlab_all_projects if self._filter_project_name == project.name]
            # 模糊匹配
            # self._filter_projects = [project for project in gitlab_all_projects if self._filter_project_name in project.name]

    def _get_project_branches(self, project):
        """
        获取项目所有分支列表或是指定分支,如果self._filter_branch不为空，获取指定的分支，否则获取所有分支
        Args:
            project: 需要获取分支列表的项目
        Returns:
            branches(list): 该项目的分支列表
        """
        branches: list = []
        if self._filter_branch is None:
            branches = project.branches.list()
        else:
            branches.append(project.branches.get(self._filter_branch))
        print("================================================================================")
        if len(branches) == 0:
            print("{}({}) has no branch named \"{}\" , please check your input branch name from gitlab."
                  .format(project.name, project.web_url, self._filter_branch))
            return list()
        for branch in branches:
            print("{}({}) has branch {}".format(project.name, project.web_url, branch.name))
        return branches

    def calculate_gitlab_project_codes(self):
        """
        统计项目代码的入口函数
        """
        # 获取project信息
        self._get_gitlab_projects()
        if len(self._target_projects) == 0:
            print("================================================================================")
            print("there is no project named \"{}\" , please check your input project name from gitlab."
                  .format(self._filter_project_name))
            return

        # 存在clone其他项目的情况，所以通过项目名称可能会获取到多个项目地址,
        for project in self._target_projects:
            # 统计每一个分支的提交情况
            branches: list = self._get_project_branches(project)
            for branch in branches:
                gitlab_project_manager = GitLabProjectManager(project=project, branch=branch.name,
                                                              since=self._since, until=self._until)
                branch_commit_data: list = gitlab_project_manager.calculate_project_branch_commit_data()
                self._project_branch_commit_data[branch.name] = branch_commit_data

    def output_code_statistics_to_execl(self):
        """
        将各分支的提交信息保存在EXCEL表格中
        """
        workbook = Workbook()
        sheet = workbook.create_sheet('gitlab', 0)
        table_head = ['工程名称', '分支名称', '邮箱', '提交次数', '新增代码', '删除代码', '总计代码']
        for i in range(0, len(table_head)):
            sheet.cell(1, i+1).value = table_head[i]
        table_row: int = 2

        for branch_commit_data in self._project_branch_commit_data.values():
            if len(branch_commit_data) <= 0:
                continue
            for author_commit in branch_commit_data:
                sheet.cell(table_row, 1).value = author_commit['project']
                sheet.cell(table_row, 2).value = author_commit['branch']
                sheet.cell(table_row, 3).value = author_commit['author_email']
                sheet.cell(table_row, 4).value = author_commit["commits"]
                sheet.cell(table_row, 5).value = author_commit["add"]
                sheet.cell(table_row, 6).value = author_commit["del"]
                sheet.cell(table_row, 7).value = author_commit["total"]
                table_row += 1

        out_put_file: str = datetime.now().strftime('%Y%m%d%H%M%S') + '_'
        if self._since and self._until:
            out_put_file += 'git_code_' + self._filter_project_name + '_' + self._since + '_' + self._until + '.xlsx'
        else:
            out_put_file += 'git_code_' + self._filter_project_name + '.xlsx'

        workbook.save(out_put_file)
        print("================================================================================")
        print("finish counting the gitlab code, and save excel file at", os.path.abspath(out_put_file))


def command_parse():
    parse = argparse.ArgumentParser(add_help=False,
                                    description='A python script for counting gitlab code by xiaozhiguo@uniontech.com.')
    parse.add_argument('-v', '--version', action='version', version='1.3.1')
    parse.add_argument('-h', '--help', action='help')
    parse.add_argument('-p', '--project', type=str, dest='project', required=True, default=None,
                       help='One gitlab project you want to count.')
    parse.add_argument('-b', '--branch', type=str, dest='branch', required=False, default=None,
                       help='Gitlab branch you want to count, default is master.')
    parse.add_argument('--since', type=str, dest='since', required=False,
                       help='Date to start counting, just like YYYY-MM-DD, default is at the begin of project built.')
    parse.add_argument('--until', type=str, dest='until', required=False,
                       help='Date to stop counting, excluding the deadline, just like YYYY-MM-DD, default is today.')
    command_args: dict = vars(parse.parse_args())

    # command_args['project'] = 'manual'
    # command_args['since'] = '2020-10-01'
    # command_args['until'] = '2020-11-01'
    # command_args['branch'] = 'maintain/sp3_5.7'

    # 判断开始统计日期参数格式是否正确
    if command_args['since']:
        try:
            date.fromisoformat(command_args['since'])
        except (TypeError, ValueError):
            print("option since {} is invalid parameters, the format just like YYYY-MM-DD, default value will be set."
                  .format(command_args['since']))
            command_args['since'] = None
    # 判断结束统计日期参数格式是否正确，如果不正确则结束时间设置为今天
    if command_args['until']:
        try:
            date.fromisoformat(command_args['until'])
        except (TypeError, ValueError):
            print("option until {} is invalid parameters, the format just like YYYY-MM-DD, default value will be set."
                  .format(command_args['until']))
            command_args['until'] = date.today().isoformat()
    else:
        command_args['until'] = date.today().isoformat()
    return command_args


if __name__ == '__main__':
    args = command_parse()
    code_stats = GitLabProjectStats(project=args['project'], branch=args['branch'], since=args['since'], until=args['until'])
    code_stats.calculate_gitlab_project_codes()
    code_stats.output_code_statistics_to_execl()
