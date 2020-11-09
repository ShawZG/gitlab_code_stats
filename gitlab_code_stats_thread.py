# -*- coding: utf-8 -*-
#
# Copyright (C) 2020 xiao zhiguo <xiaozhiguo@uniontech.com>
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Lesser General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Lesser General Public License for more details.
#
# You should have received a copy of the GNU Lesser General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.

import os
import argparse
from datetime import date
from concurrent.futures import ThreadPoolExecutor
import gitlab
from openpyxl import Workbook


class GitLabCodeStats:
    def __init__(self, branch: str = 'master', since: str = None, until: str = None):
        self._gitlab_url: str = "https://gitlabnj.uniontech.com"
        self._gitlab_access_token: str = "DtkhMpC4zi1k1WLzw92B"
        self._branch: str = branch
        self._since: str = since
        self._until: str = until
        self._projects: list = []
        self._project_commit_codes: dict = {}
        self._file: str = 'git_code_' + self._branch + '.xlsx'
        if self._since and self._until:
            self._file = 'git_code_' + self._branch + '_' + self._since + '_' + self._until + '.xlsx'

    def _get_gitlab_projects(self):
        gl = gitlab.Gitlab(url=self._gitlab_url, private_token=self._gitlab_access_token)
        gl.auth()
        self._projects = gl.projects.list(all=True)
        # for project in self._projects:
        #     print("project id:", project.id, project.name, project.web_url)

    def _get_project_commit(self, project):
        print("get commits for", project.name)
        # if (not self._since) and (not self._until):
        commits = project.commits.list(all=True, query_parameters={'since': self._since, 'until': self._until,
                                                                   'ref_name': self._branch})
        # else:
        #     commits = project.commits.list(all=True, query_parameters={'ref_name': 'dev'})

        # 该项目所有提交代码者的邮箱
        project_committers: set = set()
        # 该项目所有提交代码的数据
        project_committer_codes: list = []
        for commit in commits:
            if ('Merge' in commit.title) or ('Merge' in commit.message):
                # print('ignore merge operation by', commit.committer_email)
                continue
            elif not ("@uniontech.com" in commit.committer_email):
                # print('ignore commit from author', commit.committer_email)
                continue
            else:
                commit_detail = project.commits.get(commit.id)
                print(commit_detail)
                # author_code_stats 保存提交代码信息
                # committed_date 提交日期, additions 新增代码, deletions 删除代码, total 总计, email 作者邮箱.
                author_code_stats = commit_detail.stats
                author_code_stats['committed_date'] = commit_detail.committed_date
                # committer_email 和 author_email 存在不一致的情况，以author_email为准。
                author_code_stats['email'] = commit.author_email
                project_committers.add(commit.committer_email)
                project_committer_codes.append(author_code_stats)
                print(author_code_stats)
        self._calculate_committer_codes(project.web_url, project_committers, project_committer_codes)

    def _get_project_commits(self):
        with ThreadPoolExecutor(max_workers=None) as executor:
            executor.map(self._get_project_commit, self._projects)

    def _calculate_committer_codes(self, project_name: str, committers: set, committer_codes: list):
        committer_statistics_codes: dict = {}
        for committer in committers:
            committer_statistics_codes[committer] = {"commits": 0, "additions": 0, "deletions": 0, "total": 0}
        for codes in committer_codes:
            author = codes["email"]
            committer_statistics_codes[author]["commits"] += 1
            committer_statistics_codes[author]["additions"] += codes["additions"]
            committer_statistics_codes[author]["deletions"] += codes["deletions"]
            committer_statistics_codes[author]["total"] += codes["total"]
        print(project_name, committer_statistics_codes)
        self._project_commit_codes[project_name] = committer_statistics_codes

    def calculate_gitlab_committer_codes(self):
        self._get_gitlab_projects()
        self._get_project_commits()

    def output_code_statistics_to_execl(self):
        # pass
        workbook = Workbook()
        sheet = workbook.create_sheet('gitlab', 0)
        table_head = ['工程名称', '分支名称', '邮箱', '提交次数', '新增代码', '删除代码', '总计代码']
        for i in range(0, len(table_head)):
            sheet.cell(1, i+1).value = table_head[i]
        table_row: int = 2
        for project_name in self._project_commit_codes:
            committer_statistics_codes: dict = self._project_commit_codes.get(project_name)
            for committer in committer_statistics_codes:
                sheet.cell(table_row, 1).value = project_name
                sheet.cell(table_row, 2).value = self._branch
                sheet.cell(table_row, 3).value = committer
                sheet.cell(table_row, 4).value = committer_statistics_codes[committer]["commits"]
                sheet.cell(table_row, 5).value = committer_statistics_codes[committer]["additions"]
                sheet.cell(table_row, 6).value = committer_statistics_codes[committer]["deletions"]
                sheet.cell(table_row, 7).value = committer_statistics_codes[committer]["total"]
                table_row += 1
        workbook.save(self._file)
        print("")
        print("finish counting the gitlab code, and save excel file at", os.path.abspath(self._file))


def command_parse():
    parse = argparse.ArgumentParser(add_help=False,
                                    description='A python script for counting gitlab code by xiaozhiguo@uniontech.com.')
    parse.add_argument('-v', '--version', action='version', version='1.0.0')
    parse.add_argument('-h', '--help', action='help')
    parse.add_argument('-b', '--branch', type=str, dest='branch', required=False, default='master',
                       help='Gitlab branch you want to count, default is master.')
    parse.add_argument('--since', type=str, dest='since', required=False,
                       help='Date to start counting, just like YYYY-MM-DD, default is at the begin of project built.')
    parse.add_argument('--until', type=str, dest='until', required=False,
                       help='Date to stop counting, just like YYYY-MM-DD, default is today.')
    command_args: dict = vars(parse.parse_args())
    command_args['since'] = '2020-11-01'
    # 判断开始统计日期参数格式是否正确
    if command_args['since']:
        try:
            date.fromisoformat(command_args['since'])
        except (TypeError, ValueError):
            print("option since {} is invalid parameters, the format just like YYYY-MM-DD, default value will be set.",
                  command_args['since'])
            command_args['since'] = None
    # 判断结束统计日期参数格式是否正确，如果不正确则结束时间设置为今天
    if command_args['until']:
        try:
            date.fromisoformat(command_args['until'])
        except (TypeError, ValueError):
            print("option until {} is invalid parameters, the format just like YYYY-MM-DD, default value will be set.",
                  command_args['until'])
            command_args['until'] = date.today().isoformat()
    else:
        command_args['until'] = date.today().isoformat()
    return command_args


if __name__ == '__main__':
    args = command_parse()
    code_stats = GitLabCodeStats(branch=args['branch'], since=args['since'], until=args['until'])
    code_stats.calculate_gitlab_committer_codes()
    code_stats.output_code_statistics_to_execl()
